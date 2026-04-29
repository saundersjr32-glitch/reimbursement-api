from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
import base64
import os
import json
from datetime import date

app = Flask(__name__)
CORS(app)  # Allow requests from Netlify

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "Reimbursement_Request_Form.xlsx")

@app.route("/", methods=["GET"])
def health():
    return jsonify({"status": "S&I Reimbursement API is running"})

@app.route("/fill-form", methods=["POST"])
def fill_form():
    try:
        data     = request.get_json()
        profile  = data["profile"]
        trips    = data["trips"]
        month    = data["month"]
        year     = data["year"]

        # Load the real template (preserves ALL formatting)
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb["Reimbursement Request"]
        ws_ex = wb["Additional Varied Travel"]

        # ── Header fields ──────────────────────────────────────────────────
        ws["E4"]  = profile.get("name", "")
        ws["O4"]  = profile.get("id", "")
        ws["T4"]  = float(profile.get("rate", 0))
        ws["E6"]  = profile.get("address", "")
        ws["T6"]  = profile.get("deptId", "")
        ws["E8"]  = profile.get("primaryAssignment", "")

        # Today's date in the requester date field
        ws["J68"] = date.today().strftime("%B %d, %Y")

        # ── Varied Travel rows 31–49 (main sheet, 19 rows) ─────────────────
        main_trips  = trips[:19]
        extra_trips = trips[19:]

        for i, t in enumerate(main_trips):
            row = 31 + i
            if t.get("date"):    ws.cell(row=row, column=1).value  = t["date"]
            if t.get("from"):    ws.cell(row=row, column=3).value  = t["from"]
            if t.get("to"):      ws.cell(row=row, column=8).value  = t["to"]
            if t.get("purpose"): ws.cell(row=row, column=13).value = t["purpose"]
            if t.get("miles"):   ws.cell(row=row, column=19).value = float(t["miles"])
            if t.get("tolls"):   ws.cell(row=row, column=20).value = float(t["tolls"])

        # ── Additional Varied Travel sheet (rows 3–56) ─────────────────────
        for i, t in enumerate(extra_trips):
            row = 3 + i
            if row > 56:
                break
            if t.get("date"):    ws_ex.cell(row=row, column=1).value = t["date"]
            if t.get("from"):    ws_ex.cell(row=row, column=2).value = t["from"]
            if t.get("to"):      ws_ex.cell(row=row, column=3).value = t["to"]
            if t.get("purpose"): ws_ex.cell(row=row, column=4).value = t["purpose"]
            if t.get("miles"):   ws_ex.cell(row=row, column=5).value = float(t["miles"])
            if t.get("tolls"):   ws_ex.cell(row=row, column=6).value = float(t["tolls"])

        # ── Signature image ────────────────────────────────────────────────
        sig = profile.get("signature", "")
        if sig and "," in sig:
            sig_bytes = base64.b64decode(sig.split(",")[1])
            img = XLImage(BytesIO(sig_bytes))
            img.width  = 200
            img.height = 40
            img.anchor = "A68"
            ws.add_image(img)

        # ── Save to memory and return ──────────────────────────────────────
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        safe_name = profile.get("name", "Employee").replace(" ", "_")
        filename  = f"{safe_name}_{month}{year}_Reimbursement.xlsx"

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
